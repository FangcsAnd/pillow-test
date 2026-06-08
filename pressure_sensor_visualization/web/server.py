import sys
import os
import struct
import numpy as np
import serial
import serial.tools.list_ports
import csv
import threading
import json
import openpyxl
from datetime import datetime
from flask import Flask, render_template, request, jsonify, make_response
from flask_socketio import SocketIO, emit

SENSOR_SIZE = 16
FRAME_HEADER = bytes([0x55, 0xAA])
FRAME_TAIL = 0x5A

app = Flask(__name__)
app.config['SECRET_KEY'] = 'pillow_test_secret'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

serial_thread = None
serial_lock = threading.Lock()
serial_running = False
serial_obj = None
current_data = np.zeros((SENSOR_SIZE, SENSOR_SIZE), dtype=np.uint8).tolist()
recording = False
record_samples = []

def calculate_params(data):
    d = np.array(data, dtype=np.uint8)
    max_val = int(d.max())
    avg_val = int(d.mean())
    cp = np.sum(d > 10)
    contact_area = int(cp / (SENSOR_SIZE * SENSOR_SIZE) * 100)
    std_val = int(d.std())

    peak_pos = np.unravel_index(d.argmax(), d.shape)
    peak_pos_str = f"({peak_pos[1]},{peak_pos[0]})"
    peak_val = max_val

    flat = d.flatten()
    flat = flat[flat > 0]
    if len(flat) > 0:
        concentration = int((flat ** 2).sum() / max((flat.sum() ** 2), 1) * 100)
        srt = np.sort(flat)[::-1]
        p95_count = max(1, int(len(srt) * 0.05))
        p95_val = int(srt[:p95_count].mean())
    else:
        concentration = 0
        p95_val = 0

    gradient_sum = 0
    gradient_count = 0
    for i in range(SENSOR_SIZE):
        for j in range(SENSOR_SIZE - 1):
            gradient_sum += abs(int(d[i, j]) - int(d[i, j + 1]))
            gradient_count += 1
        if i < SENSOR_SIZE - 1:
            for j in range(SENSOR_SIZE):
                gradient_sum += abs(int(d[i, j]) - int(d[i + 1, j]))
                gradient_count += 1
    gradient = int(gradient_sum / max(gradient_count, 1))

    head_region = d[:10, :].sum()
    neck_region = d[10:13, :].sum()
    shoulder_region = d[13:, :].sum()
    total = head_region + neck_region + shoulder_region
    if total > 0:
        head_ratio = int(head_region / total * 100)
        neck_ratio = int(neck_region / total * 100)
        shoulder_ratio = int(shoulder_region / total * 100)
    else:
        head_ratio = neck_ratio = shoulder_ratio = 0

    neck_continuity = int(d[10:13, :].std())

    y_coords, x_coords = np.where(d > 10)
    if len(x_coords) > 0:
        cx = int(x_coords.mean() / SENSOR_SIZE * 100)
        cy = int(y_coords.mean() / SENSOR_SIZE * 100)
        center = f"{cx}%,{cy}%"
    else:
        center = "0%,0%"

    neck_gap = "无"
    if len(x_coords) >= 4:
        mid_row = d[7, :]
        bottom_row = d[15, :]
        if mid_row.mean() < bottom_row.mean() * 0.5:
            neck_gap = "有"

    return {
        "max": int(max_val), "avg": int(avg_val), "area": int(contact_area), "std": int(std_val),
        "peak_pos": str(peak_pos_str), "peak_val": int(peak_val), "concentration": int(concentration),
        "p95": int(p95_val), "gradient": int(gradient), "head_ratio": int(head_ratio),
        "neck_ratio": int(neck_ratio), "shoulder_ratio": int(shoulder_ratio),
        "neck_continuity": int(neck_continuity), "neck_gap": str(neck_gap), "center": str(center)
    }


def serial_read_loop(port, baudrate):
    global current_data, recording, record_samples, serial_running, serial_obj
    serial_running = True
    try:
        ser = serial.Serial(port, baudrate, timeout=0.5)
        serial_obj = ser
        buf = b''
        while serial_running:
            if ser.in_waiting > 0:
                d = ser.read(ser.in_waiting)
                buf += d
                while len(buf) >= 263:
                    idx = buf.find(FRAME_HEADER)
                    if idx == -1:
                        buf = b''
                        break
                    buf = buf[idx:]
                    if len(buf) >= 263:
                        length = struct.unpack('<H', buf[2:4])[0]
                        if length == 257 and buf[4] == 0x01 and buf[262] == FRAME_TAIL:
                            pressure = np.frombuffer(buf[5:261], dtype=np.uint8)
                            pressure = pressure.reshape((SENSOR_SIZE, SENSOR_SIZE), order='F')
                            data_list = pressure.tolist()
                            current_data = data_list
                            if recording:
                                record_samples.append(data_list)
                            params = calculate_params(data_list)
                            socketio.emit('sensor_data', {'grid': data_list, 'params': params})
                            buf = buf[263:]
                        else:
                            buf = buf[1:]
        if ser and ser.is_open:
            ser.close()
    except Exception as e:
        print(f"Serial error: {e}")
        socketio.emit('serial_error', {'message': str(e)})
    finally:
        serial_running = False
        serial_obj = None


@app.route('/')
def index():
    resp = make_response(render_template('index.html'))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    return resp


@app.route('/api/ports')
def get_ports():
    ports = [p.device for p in serial.tools.list_ports.comports()]
    return jsonify({'ports': ports})


@app.route('/api/connect', methods=['POST'])
def connect_serial():
    global serial_thread
    data = request.json
    port = data.get('port', '')
    baudrate = int(data.get('baudrate', 460800))
    with serial_lock:
        if serial_thread and serial_thread.is_alive():
            return jsonify({'status': 'already_connected', 'port': port})
        serial_thread = threading.Thread(target=serial_read_loop, args=(port, baudrate), daemon=True)
        serial_thread.start()
    return jsonify({'status': 'connected', 'port': port, 'baudrate': baudrate})


@app.route('/api/disconnect', methods=['POST'])
def disconnect_serial():
    global serial_running, serial_obj, recording
    serial_running = False
    recording = False
    record_samples.clear()
    if serial_obj and serial_obj.is_open:
        serial_obj.close()
        serial_obj = None
    return jsonify({'status': 'disconnected'})


@app.route('/api/record/start', methods=['POST'])
def start_recording():
    global recording, record_samples
    recording = True
    record_samples.clear()
    return jsonify({'status': 'recording'})


@app.route('/api/record/stop', methods=['POST'])
def stop_recording():
    global recording, record_samples
    recording = False
    samples = [s for s in record_samples]
    if len(samples) < 2:
        return jsonify({'status': 'error', 'message': '采样数据不足'})

    all_params = []
    for s in samples:
        all_params.append(calculate_params(s))

    keys = all_params[0].keys()
    avg_params = {}
    for k in keys:
        vals = [p[k] for p in all_params]
        if all(isinstance(v, (int, float)) for v in vals):
            avg_params[k] = int(np.mean(vals))
        else:
            avg_params[k] = vals[-1]

    avg_grid = np.mean(np.array(samples, dtype=np.float64), axis=0).astype(np.uint8).tolist()

    return jsonify({
        'status': 'ok',
        'samples': len(samples),
        'duration': 3.0,
        'avg_params': avg_params,
        'avg_grid': avg_grid
    })


@app.route('/api/export', methods=['POST'])
def export_csv():
    records = request.json.get('records', [])
    if not records:
        return jsonify({'status': 'error', 'message': '无记录'})

    filename = f"pillow_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', filename)

    grid_fields = []
    for i in range(SENSOR_SIZE):
        for j in range(SENSOR_SIZE):
            grid_fields.append(f'g_{i}_{j}')

    fields = ['user_id', 'name', 'gender', 'age', 'height', 'weight',
              'shoulder_width', 'neck_curve', 'neck_history', 'neck_pain',
              'pillow_id', 'pillow_brand', 'pillow_material', 'pillow_size',
              'pillow_head_height', 'pillow_neck_height', 'pillow_side_height',
              'pillow_center_hardness', 'pillow_neck_hardness',
              'pillow_edge_hardness_l', 'pillow_edge_hardness_r',
              'sleep_pos', 'comfort_total',
              'comfort_overall', 'comfort_neck', 'comfort_head',
              'comfort_supine', 'comfort_side', 'comfort_even',
              'comfort_fall', 'comfort_turn', 'comfort_pressure', 'comfort_heat',
              'time', 'note', 'samples', 'duration_s',
              'avg_max', 'avg_avg', 'avg_area', 'avg_std', 'avg_peak_pos',
              'avg_peak_val', 'avg_concentration', 'avg_p95', 'avg_gradient',
              'avg_head_ratio', 'avg_neck_ratio', 'avg_shoulder_ratio',
              'avg_neck_continuity', 'avg_neck_gap', 'avg_center'] + grid_fields

    try:
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
            writer.writeheader()
            base_count = len(fields) - len(grid_fields)
            for r in records:
                row = {k: r.get(k, '') for k in fields[:base_count]}
                comfort = r.get('comfort', {})
                if isinstance(comfort, dict):
                    for k, v in comfort.items():
                        kk = 'comfort_' + k if not k.startswith('comfort_') else k
                        if kk in row:
                            row[kk] = v
                else:
                    row['comfort_total'] = comfort
                avg_grid = r.get('avg_grid', '[]')
                try:
                    grid = json.loads(avg_grid) if isinstance(avg_grid, str) else avg_grid
                    for i in range(SENSOR_SIZE):
                        for j in range(SENSOR_SIZE):
                            row[f'g_{i}_{j}'] = int(grid[i][j]) if i < len(grid) and j < len(grid[i]) else 0
                except:
                    pass
                writer.writerow(row)
        return jsonify({'status': 'ok', 'filename': filename, 'count': len(records)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


def load_pillow_data():
    xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', '枕头舒适度数据采集表_260605.xlsx')
    pillows = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0]:
                continue
            pillows.append({
                'id': str(row[0]),
                'brand': str(row[1]) if row[1] else '',
                'material': str(row[2]) if row[2] else '',
                'center_hardness': row[3] if row[3] else '',
                'neck_hardness': row[4] if row[4] else '',
                'edge_hardness_l': row[5] if row[5] else '',
                'edge_hardness_r': row[6] if row[6] else '',
                'head_height': row[7] if row[7] else '',
                'neck_height': row[8] if row[8] else '',
                'side_height': row[9] if row[9] else '',
                'size': str(row[11]) if row[11] else '',
            })
    except Exception as e:
        print(f"Error loading pillow data: {e}")
    return pillows


@app.route('/api/pillows')
def get_pillows():
    pillows = load_pillow_data()
    return jsonify({'pillows': pillows})


def load_user_data():
    xlsx_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', '试信息库_260605.xlsx')
    users = []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[0]:
                continue
            users.append({
                'id': str(row[0]),
                'name': str(row[1]).strip() if row[1] else '',
                'gender': str(row[2]) if row[2] else '',
                'age': row[3] if row[3] else '',
                'height': row[4] if row[4] else '',
                'weight': row[5] if row[5] else '',
                'shoulder_width': row[6] if row[6] else '',
                'neck_curve': str(row[7]) if row[7] else '',
                'neck_history': str(row[8]) if row[8] else '',
                'neck_pain': str(row[9]) if row[9] else '',
            })
    except Exception as e:
        print(f"Error loading user data: {e}")
    return users


@app.route('/api/users')
def get_users():
    users = load_user_data()
    return jsonify({'users': users})


RECORDS_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), 'data', 'test_records.json')


@app.route('/api/save', methods=['POST'])
def save_records():
    records = request.json.get('records', [])
    if not records:
        return jsonify({'status': 'error', 'message': '无记录'})
    try:
        existing = []
        if os.path.exists(RECORDS_FILE):
            with open(RECORDS_FILE, 'r', encoding='utf-8') as f:
                existing = json.load(f)
        existing.extend(records)
        with open(RECORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
        return jsonify({'status': 'ok', 'count': len(records), 'total': len(existing)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/history')
def get_history():
    if not os.path.exists(RECORDS_FILE):
        return jsonify({'history': []})
    with open(RECORDS_FILE, 'r', encoding='utf-8') as f:
        history = json.load(f)
    return jsonify({'history': history})


def handle_params_request():
    if current_data:
        params = calculate_params(current_data)
        emit('params_update', params)


if __name__ == '__main__':
    print("Starting pillow test web server at http://localhost:8080")
    socketio.run(app, host='0.0.0.0', port=8080, debug=False, allow_unsafe_werkzeug=True)
